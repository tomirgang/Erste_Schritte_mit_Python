from openpyxl import load_workbook

wb = load_workbook(filename = 'Employee Sample Data.xlsx')
ws = wb.active
i = 2
sum = 0
count = 0
while ws[f'A{i}'].value:
    if ws[f'N{i}'].value == '':
        count = count + 1
        # print(f"{i}: {ws[f'B{i}'].value} verdient {'${:0,.2f}'.format(ws[f'J{i}'].value)}")
        sum = sum + ws[f'J{i}'].value
        # print("Durchschnittsverdienst: ${:0,.2f}".format(sum/count))
    i = i + 1
print("Durchschnittsverdienst: ${:0,.2f}".format(sum/count))

# Schreibe die Namen und Berufstitel aller aktiver Angestellter in ein LibreOffice Calc Dokument.

from collections import OrderedDict
from pyexcel_ods import save_data

zeilen = [['Name', 'Berufstitel']]
i = 2
while ws[f'A{i}'].value:
    if ws[f'N{i}'].value == '':
        zeilen.append([ws[f'B{i}'].value, ws[f'C{i}'].value])
    i = i + 1

data = OrderedDict()
data.update({"Angestellte": zeilen})
save_data("angestellte.ods", data)
